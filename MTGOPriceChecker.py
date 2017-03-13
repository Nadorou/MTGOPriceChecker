from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook

# TODO GUI with ability to specify path/update prices
# TODO Add ability to view entries through GUI

def getCardPrice(cardName, cardSet):
    formatted_name = cardName.replace(' ', '+')
    formatted_set = cardSet.replace(' ', '+')
    target_url = 'https://www.mtggoldfish.com/price/' + formatted_set + '/' + formatted_name + '#online'
    target_page = request_page(target_url)
    price = priceScan(target_page.text)
    return price


def request_page(url):
    page = requests.get(url)
    return page


def priceScan(htmlString):
    souped_page = BeautifulSoup(htmlString, 'html.parser')
    scanned_price = souped_page.find(class_='price-box online').find(class_='price-box-price').string
    price = float(scanned_price)
    return price


def load_excel_data(path):
    investment_wb = load_workbook(path)
    investment_ws = investment_wb['Blad1']
    card_registry = {}


    heading_index = 0
    for heading in investment_ws.iter_cols(min_row=1, max_row=1, min_col=1):
        label = heading[0].value
        if label is None:
            break
        elif label == 'Card':
            card_index = heading_index
        elif label == 'Set':
            set_index = heading_index
        elif label == 'Selling price':
            price_index = heading_index
        heading_index += 1

    for entry in investment_ws.iter_rows(min_row=2):
        if entry[card_index].value is None:
            break
        else:
            card_name = entry[card_index].value
            set_name = entry[set_index].value
            card_id = card_name + set_name
            if card_id not in card_registry.keys():
                card_price = getCardPrice(card_name, set_name)
                card_registry[card_id] = card_price

                entry[price_index].value = card_price
            else:
                entry[price_index].value = card_registry[card_id]
    investment_wb.save(path)

load_excel_data('MTGInvestments.xlsx')